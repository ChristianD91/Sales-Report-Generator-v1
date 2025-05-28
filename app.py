from pathlib import Path
import json

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import tkinter as tk
from tkinter import filedialog, messagebox
import threading
from fpdf import FPDF
import pandas as pd

SETTINGS_FILE = "settings.json"


# PyInstaller compatibility things
def resource_path(relative_path: str) -> str:
    """
    :param relative_path: str; path to resource relative to current working directory
    Get absolute path to resource, works for dev and for PyInstaller.
    """
    base_path = Path(__file__).parent  # When bundled by PyInstaller
    out_path = base_path / relative_path
    return out_path


# Reporting Logic
def generate_report(
    input_path: str, output_folder: str, selected_reports: list, output_format: str
):
    """
    :param input_path: str; path to excel file
    :param output_folder: str; path to output folder
    :param selected_reports: list; list of reports to generate
    :param output_format: str; output format

    Required fields:
    ORDERDATE
    PRODUCTLINE
    SALES
    CUSTOMERNAME
    COUNTRY
    """
    output_folder_path = Path(output_folder)
    output_folder_path.mkdir(parents=True, exist_ok=True)
    ext = input_path[input_path.rfind(".") + 1 :].lower()

    def read_csv_data(filename: str):
        with open(filename, "r") as f:
            temp = f.readline()
        first_line = temp[: temp.find("\n")]
        common_delim = ["\t", ",", "|"]
        separator = ""
        no_cols = 1
        filename = filename
        for delim in common_delim:
            sep_ct = first_line.count(delim)
            if (sep_ct + 1) > no_cols:
                separator = delim
                no_cols = sep_ct + 1
        if no_cols == 1:
            messagebox.showerror(
                title="Unknown Delimiter", message="Delimiter cannot be determined"
            )
        out_df = pd.read_csv(filename, sep=separator)
        return out_df

    if ext == ".csv":
        df = read_csv_data(input_path)
    elif ext in [".xlsx", ".xls"]:
        df = pd.read_excel(input_path)
    else:
        messagebox.showerror(
            title="Invalid File", message="Unsupported file format selected."
        )
        return

    df.columns = df.columns.str.strip()
    total_sales = df["SALES"].sum()
    df["ORDERDATE"] = pd.to_datetime(df["ORDERDATE"], errors="coerce")
    df["Month"] = df["ORDERDATE"].dt.to_period("M")
    df["Quarter"] = df["ORDERDATE"].dt.to_period("Q")

    wb = Workbook()
    summary_data = []

    def add_df_to_sheet(sheet_name, data_frame, chart_column=None):
        ws = wb.create_sheet(title=sheet_name)
        for col_num, column_title in enumerate(data_frame.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = Font(bold=True)

        for row_num, row in enumerate(data_frame.itertuples(index=False), 2):
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                if col_num == 2 and isinstance(value, (int, float)):
                    cell.number_format = '"$"#,##0.00'

        for i in range(1, len(data_frame.columns) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 30

        # Optional chart
        if chart_column is not None:
            chart = BarChart()
            chart.title = sheet_name
            chart.y_axis.title = "Sales"
            chart.x_axis.title = data_frame.columns[0]

            data = Reference(ws, min_col=2, min_row=1, max_row=len(data_frame) + 1)
            categories = Reference(
                ws, min_col=1, min_row=2, max_row=len(data_frame) + 1
            )
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            ws.add_chart(chart, f"E2")

    if "Top Products" in selected_reports:
        df_top_products = (
            df.groupby("PRODUCTLINE")["SALES"]
            .sum()
            .sort_values(ascending=False)
            .head(5)
            .reset_index()
        )
        add_df_to_sheet("Top Products", df_top_products, chart_column="SALES")
        summary_data.append(("Top Products", df_top_products))

    if "Top Customers" in selected_reports:
        df_top_customers = (
            df.groupby("CUSTOMERNAME")["SALES"]
            .sum()
            .sort_values(ascending=False)
            .head(5)
            .reset_index()
        )
        add_df_to_sheet("Top Customers", df_top_customers, chart_column="SALES")
        summary_data.append(("Top Customers", df_top_customers))

    if "Sales by Country" in selected_reports:
        df_country = (
            df.groupby("COUNTRY")["SALES"]
            .sum()
            .sort_values(ascending=False)
            .reset_index()
        )
        add_df_to_sheet("Sales by Country", df_country, chart_column="SALES")
        summary_data.append(("Sales by Country", df_country))

    if "Sales by Month" in selected_reports:
        df_month = df.groupby("Month")["SALES"].sum().reset_index()
        df_month["Month"] = df_month["Month"].astype(str)
        add_df_to_sheet("Sales by Month", df_month, chart_column="SALES")
        summary_data.append(("Sales by Month", df_month))

    if "Sales by Quarter" in selected_reports:
        df_quarter = df.groupby("Quarter")["SALES"].sum().reset_index()
        df_quarter["Quarter"] = df_quarter["Quarter"].astype(str)
        add_df_to_sheet("Sales by Quarter", df_quarter, chart_column="SALES")
        summary_data.append(("Sales by Quarter", df_quarter))

    if "Sales by Range" in selected_reports:
        min_sale, max_sale = df["SALES"].min(), df["SALES"].max()
        low_thresh = min_sale + (max_sale - min_sale) / 3
        high_thresh = min_sale + 2 * (max_sale - min_sale) / 3

        def range_category(s):
            if s <= low_thresh:
                return "Low Range"
            elif s <= high_thresh:
                return "Middle Range"
            else:
                return "High Range"

        df["Range Category"] = df["SALES"].apply(range_category)
        df_range = (
            df.groupby("Range Category")["SALES"]
            .sum()
            .reindex(["Low Range", "Middle Range", "High Range"])
            .reset_index()
        )
        add_df_to_sheet("Sales by Range", df_range, chart_column="SALES")
        summary_data.append(("Sales by Range", df_range))

    if "Total Sales Summary" in selected_reports:
        ws = wb.active
        ws.title = "Summary"
        ws["A1"] = "Total Sales"
        ws["A1"].font = Font(bold=True)
        ws["B1"] = total_sales
        ws["B1"].number_format = '"$"#,##0.00'
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 30

    filename = output_folder_path / f"sales_summary_report.{output_format}"

    if output_format == "xlsx":
        wb.save(filename)
    elif output_format == "csv":
        if summary_data:
            summary_data[0][1].to_csv(filename, index=False)
    elif output_format == "pdf":
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        pdf.cell(200, 10, txt="Sales Summary Report", ln=True, align="C")
        for title, data in summary_data:
            pdf.ln(10)
            pdf.set_font("Arial", "B", 12)
            pdf.cell(200, 10, title, ln=True)
            pdf.set_font("Arial", size=10)
            for row in data.itertuples(index=False):
                row_text = ", ".join(str(val) for val in row)
                pdf.multi_cell(0, 8, row_text)
        pdf.output(filename)

    messagebox.showinfo("Report Generated", f"✅ Report saved to: {filename}")

    # Saves previous paths & report selections
    with open(SETTINGS_FILE, "w") as f:
        json.dump(
            {
                "last_input": input_path,
                "last_output": output_folder,
                "last_format": output_format,
                "last_options": selected_reports,
            },
            f,
        )


# ---------- GUI ----------
def launch_gui():
    window = tk.Tk()
    window.iconbitmap(resource_path("sales_report_generator.ico"))
    window.title("Sales Report Generator")
    window.geometry("600x500")

    file_path = tk.StringVar()
    output_path = tk.StringVar()
    theme_var = tk.StringVar(value="Default")
    format_var = tk.StringVar(value="xlsx")

    options = [
        "Top Products",
        "Top Customers",
        "Sales by Country",
        "Sales by Month",
        "Total Sales Summary",
        "Sales by Quarter",
        "Sales by Range",
    ]
    checks = {}

    settings_file = Path(SETTINGS_FILE)
    if settings_file.is_file():
        try:
            with open(SETTINGS_FILE, "r") as f:
                saved = json.load(f)
                file_path.set(saved.get("last_input", ""))
                output_path.set(saved.get("last_output", ""))
                format_var.set(saved.get("last_format", "xlsx"))
        except Exception:
            pass

    # Layout
    top_frame = tk.Frame(window)
    top_frame.pack(pady=10)

    tk.Label(top_frame, text="Input File:").grid(row=0, column=0, sticky="w")
    tk.Entry(top_frame, textvariable=file_path, width=35).grid(row=0, column=1)
    tk.Button(
        top_frame,
        text="Browse",
        command=lambda: file_path.set(
            filedialog.askopenfilename(filetypes=[("Data Files", "*.csv *.xlsx *.xls")])
        ),
    ).grid(row=0, column=2)

    tk.Label(top_frame, text="Output Folder:").grid(row=1, column=0, sticky="w")
    tk.Entry(top_frame, textvariable=output_path, width=35).grid(row=1, column=1)
    tk.Button(
        top_frame,
        text="Browse",
        command=lambda: output_path.set(filedialog.askdirectory()),
    ).grid(row=1, column=2)

    tk.Label(top_frame, text="Output Format:").grid(row=2, column=0, sticky="w")
    format_menu = tk.OptionMenu(top_frame, format_var, "xlsx", "csv", "pdf")
    format_menu.grid(row=2, column=1, sticky="w")

    tk.Label(top_frame, text="Theme:").grid(row=3, column=0, sticky="w")
    theme_menu = tk.OptionMenu(
        top_frame,
        theme_var,
        "Default",
        "Windows 98",
        command=lambda _: apply_theme(theme_var.get()),
    )
    theme_menu.grid(row=3, column=1, sticky="w")

    # Checkbox area
    checkbox_frame = tk.Frame(window)
    checkbox_frame.pack(pady=5)

    for i, option in enumerate(options):
        var = tk.BooleanVar(value=True)
        checks[option] = var
        tk.Checkbutton(checkbox_frame, text=option, variable=var).grid(
            row=i // 2, column=i % 2, sticky="w"
        )

    # Generate Button
    def on_generate():
        selected = [opt for opt, var in checks.items() if var.get()]
        if not file_path.get() or not output_path.get() or not selected:
            messagebox.showwarning(
                "Missing Info", "Please fill all fields and select options."
            )
            return
        threading.Thread(
            target=generate_report,
            args=(file_path.get(), output_path.get(), selected, format_var.get()),
            daemon=True,
        ).start()

    tk.Button(
        window,
        text="Generate Report",
        command=on_generate,
        bg="red",
        fg="white",
        font=("Comic Sans MS", 10, "bold"),
        relief="raised",
        borderwidth=3,
    ).pack(pady=10)

    def apply_theme(name):
        if name == "Windows 98":
            window.config(bg="#000080")
            for f in [top_frame, checkbox_frame]:
                f.config(bg="#000080")
                for child in f.winfo_children():
                    if isinstance(child, tk.Label):
                        child.config(
                            bg="#000080",
                            fg="#FFFF00",
                            font=("Comic Sans MS", 10, "bold"),
                        )
                    elif isinstance(child, tk.Entry):
                        child.config(
                            bg="#FFFF00",
                            fg="#000080",
                            font=("Comic Sans MS", 10, "bold"),
                        )
                    elif isinstance(child, (tk.Button, tk.OptionMenu)):
                        child.config(
                            bg="#FF0000",
                            fg="#FFFF00",
                            font=("Comic Sans MS", 10, "bold"),
                        )
            for child in checkbox_frame.winfo_children():
                if isinstance(child, tk.Checkbutton):
                    child.config(
                        bg="#000080",
                        fg="#FFFF00",
                        selectcolor="#000080",
                        font=("Comic Sans MS", 9),
                    )
        else:
            window.config(bg="SystemButtonFace")
            for f in [top_frame, checkbox_frame]:
                f.config(bg="SystemButtonFace")
                for child in f.winfo_children():
                    child.config(bg="SystemButtonFace", fg="black", font=("Arial", 9))
                    if isinstance(child, tk.Entry):
                        child.config(bg="white", fg="black")
                    if isinstance(child, tk.Button):
                        child.config(bg="SystemButtonFace", fg="black")
                    if isinstance(child, tk.Checkbutton):
                        child.config(selectcolor="SystemButtonFace")

    apply_theme(theme_var.get())
    window.mainloop()


if __name__ == "__main__":
    launch_gui()
