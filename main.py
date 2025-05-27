import pandas as pd
import os
import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# This generates reports based on the selected checkboxes
def generate_report(csv_path, output_folder, options):
    os.makedirs(output_folder, exist_ok=True)
    df = pd.read_csv(csv_path, encoding='latin1')
    df.columns = df.columns.str.strip()

    df['ORDERDATE'] = pd.to_datetime(df['ORDERDATE'], errors='coerce')
    df['Month'] = df['ORDERDATE'].dt.to_period('M')
    df['Quarter'] = df['ORDERDATE'].dt.to_period('Q')

    total_sales = df['SALES'].sum()
    wb = Workbook()

    def add_df_to_sheet(wb, sheet_name, df):
        ws = wb.create_sheet(title=sheet_name)
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = Font(bold=True)
        for row_num, row in enumerate(df.itertuples(index=False), 2):
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                if isinstance(value, (int, float)):
                    cell.number_format = '"$"#,##0.00'
        for col_idx in range(1, df.shape[1] + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 30

    # Saves cleaned up data
    df.to_csv(os.path.join(output_folder, "cleaned_sales_data.csv"), index=False)

    if options['top_products']:
        top_products = df.groupby('PRODUCTLINE')['SALES'].sum().sort_values(ascending=False).head(5)
        df_top_products = top_products.reset_index()
        df_top_products.to_csv(os.path.join(output_folder, "top_products.csv"), index=False)
        add_df_to_sheet(wb, "Top Products", df_top_products)

    if options['top_customers']:
        top_customers = df.groupby('CUSTOMERNAME')['SALES'].sum().sort_values(ascending=False).head(5)
        df_top_customers = top_customers.reset_index()
        df_top_customers.to_csv(os.path.join(output_folder, "top_customers.csv"), index=False)
        add_df_to_sheet(wb, "Top Customers", df_top_customers)

    if options['by_country']:
        sales_by_country = df.groupby('COUNTRY')['SALES'].sum().sort_values(ascending=False)
        df_sales_by_country = sales_by_country.reset_index()
        df_sales_by_country.to_csv(os.path.join(output_folder, "sales_by_country.csv"), index=False)
        add_df_to_sheet(wb, "Sales by Country", df_sales_by_country)

    if options['by_month']:
        sales_by_month = df.groupby('Month')['SALES'].sum()
        sales_by_month.index = sales_by_month.index.astype(str)
        df_sales_by_month = sales_by_month.reset_index()
        df_sales_by_month.to_csv(os.path.join(output_folder, "sales_by_month.csv"), index=False)
        add_df_to_sheet(wb, "Sales by Month", df_sales_by_month)

    if options['by_quarter']:
        sales_by_quarter = df.groupby('Quarter')['SALES'].sum()
        sales_by_quarter.index = sales_by_quarter.index.astype(str)
        df_sales_by_quarter = sales_by_quarter.reset_index()
        df_sales_by_quarter.to_csv(os.path.join(output_folder, "sales_by_quarter.csv"), index=False)
        add_df_to_sheet(wb, "Sales by Quarter", df_sales_by_quarter)

    if options['by_range']:
        min_val = df['SALES'].min()
        max_val = df['SALES'].max()
        range_size = (max_val - min_val) / 3

        # Defines the bins
        bins = [min_val, min_val + range_size, min_val + 2 * range_size, max_val + 1]
        labels = ['Low Range', 'Middle Range', 'High Range']

        # Assigns ranges
        df['Range'] = pd.cut(df['SALES'], bins=bins, labels=labels, include_lowest=True)

        # Sums sales per range
        sales_by_range = df.groupby('Range')['SALES'].sum().reindex(labels)

        # Extracts ranges to be displayed
        df_sales_by_range = pd.DataFrame({
            'Range': labels,
            'Minimum sale value': [bins[0], bins[1]+0.01, bins[2]+0.01],
            'Maximum sale value': [bins[1], bins[2], max_val],
            'Sales': sales_by_range.values
        })

        df_sales_by_range.to_csv(os.path.join(output_folder, "sales_by_range.csv"), index=False)
        add_df_to_sheet(wb, "Sales by Range", df_sales_by_range)

    if options['total_summary']:
        summary_ws = wb.active
        summary_ws.title = "Summary"
        summary_ws["A1"] = "Total Sales"
        summary_ws["A1"].font = Font(bold=True)
        summary_ws["B1"] = total_sales
        summary_ws["B1"].number_format = '"$"#,##0.00'
        summary_ws.column_dimensions["A"].width = 30
        summary_ws.column_dimensions["B"].width = 30

    excel_path = os.path.join(output_folder, "sales_summary_report.xlsx")
    wb.save(excel_path)
    messagebox.showinfo("Report Generated", f"Excel summary report saved to:\n{excel_path}")

# GUI
root = tk.Tk()
root.title("Sales Report Generator")
root.geometry("500x400")

csv_path = tk.StringVar()
output_folder = tk.StringVar(value="output")

options = {
    'top_products': tk.BooleanVar(value=True),
    'top_customers': tk.BooleanVar(value=True),
    'by_country': tk.BooleanVar(value=True),
    'by_month': tk.BooleanVar(value=True),
    'total_summary': tk.BooleanVar(value=True),
    'by_quarter': tk.BooleanVar(value=True),
    'by_range': tk.BooleanVar(value=True)
}

def browse_csv():
    path = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
    if path:
        csv_path.set(path)

def browse_output():
    folder = filedialog.askdirectory()
    if folder:
        output_folder.set(folder)

def run_report():
    if not csv_path.get():
        messagebox.showerror("Input Required", "Please select a CSV file.")
        return
    generate_report(csv_path.get(), output_folder.get(), {k: v.get() for k, v in options.items()})

# Layout
frame = tk.Frame(root, padx=20, pady=20)
frame.pack(fill="both", expand=True)

tk.Label(frame, text="Select CSV File:").grid(row=0, column=0, sticky="w")
tk.Entry(frame, textvariable=csv_path, width=50).grid(row=0, column=1, padx=5)
tk.Button(frame, text="Browse", command=browse_csv).grid(row=0, column=2, padx=5, pady=5)

tk.Label(frame, text="Select Output Folder:").grid(row=1, column=0, sticky="w")
tk.Entry(frame, textvariable=output_folder, width=50).grid(row=1, column=1, padx=5)
tk.Button(frame, text="Browse", command=browse_output).grid(row=1, column=2, padx=5, pady=5)

tk.Label(frame, text="Select Report Options:", font=("Arial", 10, "bold")).grid(row=2, column=0, sticky="w", pady=(10, 0), columnspan=3)

options_list = [
    ('top_products', 'Top Products'),
    ('top_customers', 'Top Customers'),
    ('by_country', 'Sales by Country'),
    ('by_month', 'Sales by Month'),
    ('total_summary', 'Total Sales Summary'),
    ('by_quarter', 'Sales by Quarter'),
    ('by_range', 'Sales by Range (Low/Mid/High)')
]

for idx, (key, label) in enumerate(options_list):
    col = idx // 4
    row = 3 + (idx % 4)
    tk.Checkbutton(frame, text=label, variable=options[key]).grid(row=row, column=col, sticky="w")

tk.Button(frame, text="Generate Report", command=run_report, bg="#4CAF50", fg="white", padx=10, pady=5).grid(row=7, column=0, columnspan=3, pady=20)

root.mainloop()
